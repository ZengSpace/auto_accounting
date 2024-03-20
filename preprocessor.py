import pandas as pd
from openpyxl import load_workbook

def preprocessor_xlsx(file_path):
    # Load the Excel workbook
    workbook = load_workbook(file_path)

    # Select the desired worksheet
    worksheet = workbook.active

    # Delete the first column
    worksheet.delete_cols(1)

    # Define the range of rows to skip (second to fifth rows and the last 8 rows)
    rows_to_skip = list(range(2, 6)) + list(range(worksheet.max_row - 7, worksheet.max_row + 1))

    # Iterate through rows to skip and remove them
    for row_index in sorted(rows_to_skip, reverse=True):
        worksheet.delete_rows(row_index)

    # Save the modified workbook
    workbook.save('BalanceSheetDetail_modified.xlsx')

# Example usage:
file_path = 'BalanceSheetDetail-404.xlsx'
statistics = preprocessor_xlsx(file_path)

# Read the Excel file
df = pd.read_excel('BalanceSheetDetail_modified.xlsx')

# Display basic statistics
print("Basic Statistics:")
print(df.describe())

# Output statistics
print("Basic Statistics:")
print(statistics)